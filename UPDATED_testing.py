import requests, json, sys, time
from pprint import pprint
import pandas as pd
import xlsxwriter
import xlrd
from openpyxl import load_workbook


# This function includes the introduction information for the API
def intro1():
    print("Welcome to the Nutritionix API Service.")
    # time.sleep(1)  # time.sleep adds time before the next line of code is printed
    print("What is your name?")
    myName = input()
    print("Hi " + myName + "!")
    # time.sleep(1)
    # print(
    #     "With this program you can either choose recipes to cook at home and eat in, or find menu items from a restaurant and dine out. "
    # )
    # time.sleep(3)
    # print(
    #     " Once you have selected your dining option, you will either get a full recipe and related nutritional information with the dine in option, or the nutritional information from a menu item at a selected restaurant."
    # )
    # time.sleep(3)
    # print("          ")
    # print(
    #     "If you decide to dine in...\n You can search pizza, for example, and have an entire recipe of how to make pizza in your kitchen from scratch!"
    # )
    # time.sleep(3)
    # print("          ")
    # print(
    #     "Or if you decide to dine out...\n  You can search Mcdonald's, for example, and select McNuggets.\n Then, information including calories, fats, cholesterol, sugar and protein for McNuggets will be given to you."
    # )
    # print("          ")
    # time.sleep(3)
    return myName


def intro2(myName):
    print("Now, " + myName + ", would you like to eat in or dine out?")
    dineInOrOut = input()
    print("---------------")
    return dineInOrOut


# print empty excel called compareSearches
def intro3():
    # writer = pandas.ExcelWriter("compareSearches.xlsx", engine="xlsxwriter")
    writer = pd.ExcelWriter(  # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
        "nutritionalDatabase.xlsx", engine="xlsxwriter"
    )
    writer.save()


# This function asks for a food item that the restaurant selected has and outputs nutrition information
def functionA(n):
    # The while loop allows for people to select more than one food item at once
    # while True:
    print(
        "please enter which item name you want to see nutrition information for from the list below or enter q(uit)"
    )
    foodItem = input()
    # if (
    #     foodItem == "q"
    # ):  # if the user enter q the program will quit out of the foodItem function
    # break
    return foodItem


# print nutrition info for indicated menu item
def functionB(n, foodItem, myList):
    for item in n:
        if foodItem == item["fields"]["item_name"]:
            selectedFood = {
                "name": item["fields"]["item_name"],
                "calories": str(item["fields"]["nf_calories"]) + " kcals",
                # "calories from fat": str(item["fields"]["nf_calories_from_fat"])
                # + " calories from fat",
                # "calories from fat": str(item["fields"]["nf_calories_from_fat"])
                # + " calories from fat",
                "total fat": str(item["fields"]["nf_total_fat"]) + " grams",
                # "saturated fat": str(item["fields"]["nf_saturated_fat"])
                # + " grams of saturated fat",
                # "saturated fat": str(item["fields"]["nf_saturated_fat"])
                # + " grams of saturated fat",
                "cholesterol": str(item["fields"]["nf_cholesterol"]) + " grams",
                "sugars": str(item["fields"]["nf_sugars"]) + " grams",
                "protein": str(item["fields"]["nf_protein"]) + " grams",
            }
            # this adds the information gathered from selectedFood to a list called myList
            # myList prints all of the food items selected with their corresponding nutritional information after the user quits out of the program
            myList.append(selectedFood)
    # for field in selectedFood:
    #     print(
    #         selectedFood[field]
    #     )  # this print statement will print the nutritional information for the food item selected
    return selectedFood, myList


# need to ask if dine in or eat out with related code
# this function asks the player which restaurants data they want to look at and returns food items sold at the restaurant selected
def functionC():
    # while (
    #     True
    # ):  # this While loop will continue to run until the player quits out of the game, the game is ended by the break in line 16
    print("Please enter a restaurant name.")
    restaurant = input(
        "Restaurant/brand name or (q)uit>"
    )  # user will input a restaurant name and the API will search relevant information from the restaurant
    # if restaurant.lower() == "q":
    #     break
    return restaurant


def functionD(restaurant):
    # This URL calls the API with the search by 'restaurant' function and names it URL.
    # When a 'restaurant' is chosen, the adjusted url returns a dictionary within a list of food item names for that specific restaurant/brand
    url = f"https://api.nutritionix.com/v1_1/search/{restaurant}?results=0:20&fields=item_name,nf_calories,nf_calories_from_fat,nf_total_fat,nf_saturated_fat,nf_cholesterol,nf_sugars,nf_protein&appId=f3b374ec&appKey=988ab0bd10dc4a886738973eec0524d1"
    response = requests.get(url)
    response.raise_for_status()  # check for errors
    # Load json data into a python variable
    nutritionData = json.loads(response.text)
    # variable 'n' looks into
    n = nutritionData["hits"]
    return n


# printing restaurant menu
def functionE(n):
    for item in n:
        print(
            item["fields"]["item_name"],
        )


# print pandas table
def functionF(n, myList, selectedFood, foodItem):
    print("Here are the nutrition facts for the food items you selected:\n")
    time.sleep(2)
    # print("---------------")
    # print(" ")
    for item in myList:

        foodData = [
            ["Item Name:", item["name"]],
            ["Calories:", item["calories"]],
            # ["Calories from Fat:", item["calories from fat"]],
            ["Total Fat:", item["total fat"]],
            # ["Saturated Fat:", item["saturated fat"]],
            ["Cholesterol:", item["cholesterol"]],
            ["Sugar:", item["sugars"]],
            ["Protein:", item["protein"]],
        ]

        headerY = [
            "  ",
            "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
        ]
        # rows
        headerX = [
            "  ",
            "  ",
            "  ",
            "  ",
            "  ",
            "  ",
            # "  ",
            # "  ",
        ]
        df = pd.DataFrame(foodData, headerX, headerY)
        print(df)
        # print(pd.DataFrame(foodData, headerX, headerY)) - REMOVE IF THE OTHER THING WORKS

        # add to excel sheet created at beginning of program
        # writer = pd.ExcelWriter(  # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
        #     "demo2.xlsx", engine="openpyxl", mode="a"
        # )
        # # try to open an existing workbook
        # writer.book = load_workbook("demo2.xlsx")
        # # # copy existing sheets
        # # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # # read existing file
        # reader = pd.read_excel("demo2.xlsx")
        # write out the new sheet, save, and close
        # df.to_excel(
        #     writer,
        #     # "demo2.xlsx",
        #     startrow=len(reader) + 1,
        #     header=False,
        #     index=False,
        #     sheet_name="Sheet1",
        # )
        # writer.save()
        # writer.close()

        with pd.ExcelWriter(
            "nutritionalDatabase.xlsx", engine="openpyxl", mode="a"
        ) as f:
            df.to_excel(f, sheet_name=foodItem, index=False, header=False)


# append_to_excel(<your_excel_path>, <new_df>, <new_sheet_name>)


def functionG(restaurant):
    print("Do you want to find the nearest " + restaurant + "?")
    searchLocation = input()
    return searchLocation


# print URL if user says yes
def functionH(searchLocation, restaurant):
    if searchLocation == "yes" or searchLocation == "y":
        # Fetch the information of that restaurant
        restaurantNew = restaurant.replace(" ", "")
        map_url = f"https://www.google.com/maps/search/?api=1&query={restaurantNew}"
        # Print message
        print("To view " + restaurant + " in Google Maps, go to " + map_url)
        print("-----------------------------------------")
        return map_url

    elif searchLocation == "no" or searchLocation == "n":
        pass


# ask which food item user wants to cook
def function1():
    print(
        "please enter what food item you would like to cook, and we will provide the recipe"
    )
    chosenfoodItem = input()
    # if recipe.lower() == "q":
    #     break
    return chosenfoodItem


# search recipes based on user input
def function2(chosenfoodItem):
    url = f"https://api.edamam.com/search?q={chosenfoodItem}&app_id=62fd6c39&app_key=5f0e27adae50eb576e4aebe320bbefe2"

    response = requests.get(url)
    response.raise_for_status()  # check for errors

    recipeData = json.loads(response.text)
    return recipeData


def function3(recipeData):
    r = recipeData["hits"]
    return r


# print options based on what user searched
def function4(r):
    for item in r:
        # print recipe name
        recipeName = item["recipe"]["label"]
        print(str(recipeName))


# ask user to select a specific recipe based on results
def function5(chosenfoodItem):
    print(
        "Please select which "
        + chosenfoodItem
        + " you are interested in cooking, and we will show you the ingredients."
    )
    chosenRecipeName = input()
    return chosenRecipeName


# print ingredients
def function6(chosenRecipeName, r):
    for item in r:
        # print(item["recipe"]["label"])
        for ingredient in item["recipe"]["ingredientLines"]:
            if chosenRecipeName == item["recipe"]["label"]:
                print("-" + ingredient)
                pass


# ask if user wants to see nutrition facts
def function7(chosenRecipeName):
    print(
        "Would you like to see the nutritional facts for "
        + chosenRecipeName
        + "? Type yes or no"
    )
    seeNutritionInfo = input()
    return seeNutritionInfo


# calories function
def function8(r, chosenRecipeName):
    for item in r:
        # print(item["recipe"]["label"])
        # for chosenFoodCalories in item["recipe"]["calories"]:
        if chosenRecipeName == item["recipe"]["label"]:
            recipeCalories = str(int(item["recipe"]["calories"]))
            # print(chosenFoodCalories + " calories")
        pass
    return recipeCalories


# total fat function
def function9(r, chosenRecipeName):
    for item in r:
        # print(item["recipe"]["label"])
        if chosenRecipeName == item["recipe"]["label"]:
            recipeTotalFat = str(
                int(item["recipe"]["totalNutrients"]["FAT"]["quantity"])
            )
            # print(totalFat + " grams of Total Fat")
        pass
    return recipeTotalFat


# cholesterol function
def function10(r, chosenRecipeName):
    for item in r:
        # print(item["recipe"]["label"])
        if chosenRecipeName == item["recipe"]["label"]:
            recipeCholesterol = str(
                int(item["recipe"]["totalNutrients"]["CHOLE"]["quantity"])
            )
            # print(totalCHOLE + " mg of Cholesterol")
        pass
    return recipeCholesterol


# sugar function
def function11(r, chosenRecipeName):
    for item in r:
        # print(item["recipe"]["label"])
        if chosenRecipeName == item["recipe"]["label"]:
            recipeSugar = str(
                int(item["recipe"]["totalNutrients"]["SUGAR"]["quantity"])
            )
            # print(totalSugar + " grams of Sugar")
        pass
    return recipeSugar


# protein function
def function12(r, chosenRecipeName):
    for item in r:
        # print(item["recipe"]["label"])
        if chosenRecipeName == item["recipe"]["label"]:
            recipeProtein = str(
                int(item["recipe"]["totalNutrients"]["PROCNT"]["quantity"])
            )
            # print(totalProtein + " grams of Protein")
        pass
    return recipeProtein


# creating dictionary for recipe nutrition info
def functionRecipeDictionary(
    chosenRecipeName,
    recipeCalories,
    recipeTotalFat,
    recipeCholesterol,
    recipeSugar,
    recipeProtein,
    recipeNutritionDictionary,
    r,
):
    for item in r:
        if chosenRecipeName == item["recipe"]["label"]:
            recipeNutritionDictionary = {
                "Item Name: ": chosenRecipeName,
                "Calories: ": recipeCalories,
                "Total Fat: ": recipeTotalFat,
                "Cholesterol: ": recipeCholesterol,
                "Sugar: ": recipeSugar,
                "Protein: ": item["recipe"]["totalNutrients"]["PROCNT"]["quantity"],
            }
            myList.append(recipeNutritionDictionary)
            return myList, recipeNutritionDictionary


# print nutrition info
def function13(
    seeNutritionInfo,
    chosenRecipeName,
    recipeCalories,
    recipeTotalFat,
    recipeCholesterol,
    recipeSugar,
    recipeProtein,
):
    if seeNutritionInfo == "yes" or seeNutritionInfo == "y":
        recipeNutritionPandas = [
            ["Item Name:", chosenRecipeName],
            ["Calories:", recipeCalories + " kcals"],
            ["Total Fat:", recipeTotalFat + " grams"],
            ["Cholesterol:", recipeCholesterol + " grams"],
            ["Sugar:", recipeSugar + " grams"],
            ["Protein:", recipeProtein + " grams"],
        ]
        # columns
        headerY = [
            "  ",
            "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
            # "  ",
        ]
        # rows
        headerX = [
            "  ",
            "  ",
            "  ",
            "  ",
            "  ",
            "  ",
            # "  ",
            # "  ",
        ]
        # print(pd.DataFrame(recipeNutritionPandas, headerX, headerY))

        df = pd.DataFrame(recipeNutritionPandas, headerX, headerY)
        print(df)
        # print(pd.DataFrame(foodData, headerX, headerY)) - REMOVE IF THE OTHER THING WORKS

        # add to excel sheet created at beginning of program
        # writer = pd.ExcelWriter(  # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
        #     "demo2.xlsx", engine="openpyxl", mode="a"
        # )
        # # try to open an existing workbook
        # writer.book = load_workbook("demo2.xlsx")
        # # # copy existing sheets
        # # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # # read existing file
        # reader = pd.read_excel("demo2.xlsx")
        # # write out the new sheet, save, and close
        # df.to_excel(
        #     writer,
        #     "demo2.xlsx",
        #     startrow=len(reader) + 1,
        #     header=False,
        #     index=False,
        # )
        # writer.save()
        # writer.close()

        with pd.ExcelWriter(
            "nutritionalDatabase.xlsx", engine="openpyxl", mode="a"
        ) as f:
            df.to_excel(f, sheet_name=chosenRecipeName, index=False, header=False)

    elif seeNutritionInfo == "no" or seeNutritionInfo == "n":
        print("No problem!")

    # pandas.DataFrame(recipeNutritionPandas).to_excel("recipeInfo.xlsx")


# def pandasComparison(myList):
#     print("Here is an excel of all the nutrition data you've viewed so far.")
#     pandas.DataFrame(myList).to_excel("finalExcel.xlsx")


def closing(myList):
    print("Thank you for using our program. Stay healthy!")
    print(
        "Be sure to view the excel sheet titled 'compareSearches' to see all the nutrition data for your searches!"
    )


# main function
userMyName = intro1()
intro3()
exploreAgain = "yes" or "y"
while exploreAgain == "yes" or exploreAgain == "y":
    userDineInorOut = intro2(userMyName)
    # mylist = []
    if (
        userDineInorOut == "out"
        or userDineInorOut == "dine out"
        or userDineInorOut == "Dine out"
    ):
        exploreRestaurant = "yes" or "y"
        while exploreRestaurant == "yes" or exploreRestaurant == "y":
            userRestaurant = functionC()
            userN = functionD(userRestaurant)
            functionE(userN)
            userFoodItem = functionA(userN)
            myList = []
            selectedFood = {}
            functionB(userN, userFoodItem, myList)
            functionF(userN, myList, selectedFood, userFoodItem)
            # fix URL
            userSearchLocation = functionG(userRestaurant)
            functionH(userSearchLocation, userRestaurant)
            print("Do you want to explore another restaurant?")
            exploreRestaurant = input()
    elif (
        userDineInorOut == "in"
        or userDineInorOut == "eat in"
        or userDineInorOut == "eat in"
    ):
        exploreRecipe = "yes" or "y"
        while exploreRecipe == "yes" or exploreRecipe == "y":
            userChosenFoodItem = function1()
            userRecipeData = function2(userChosenFoodItem)
            userR = function3(userRecipeData)
            function4(userR)
            userChosenRecipeName = function5(userChosenFoodItem)
            function6(userChosenRecipeName, userR)
            # myList = []
            recipeNutritionDictionary = {}
            userSeeNutritionInfo = function7(userChosenRecipeName)
            userRecipeCalories = function8(userR, userChosenRecipeName)
            userRecipeTotalFat = function9(userR, userChosenRecipeName)
            userRecipeCholesterol = function10(userR, userChosenRecipeName)
            userRecipeSugar = function11(userR, userChosenRecipeName)
            userRecipeProtein = function12(userR, userChosenRecipeName)
            functionRecipeDictionary(
                recipeNutritionDictionary,
                userRecipeProtein,
                userRecipeCalories,
                userRecipeTotalFat,
                userRecipeCholesterol,
                userRecipeSugar,
                userChosenRecipeName,
                userR,
            )
            function13(
                userSeeNutritionInfo,
                userChosenRecipeName,
                userRecipeCalories,
                userRecipeTotalFat,
                userRecipeCholesterol,
                userRecipeSugar,
                userRecipeProtein,
            )
            print("Do you want to explore another recipe to cook?")
            exploreRecipe = input()
    print("Do you want to explore more dining options?")
    exploreAgain = input()
closing(myList)
